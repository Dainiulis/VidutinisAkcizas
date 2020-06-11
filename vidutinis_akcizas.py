import datetime
import time
import tkinter as tk
from calendar import monthrange
from random import randint

import numpy as np
import pandas as pd
# COL NAMES
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

SAVE_TO_PICKLE = True

SHEET_LIKUTIS_PABAIGAI_AX_GRUPUOTAS = 'Likutis pabaigai AX grupuotas'
SHEET_TARIFINES_GRUPES = 'tarifinės grupės'
SHEET_PRITRAUKTI_DUOMENYS = "Pritraukti duomenys"
SHEET_SUVESTINE = 'Suvestine'
SHEET_AKCIZAS = 'Vidutinis_akcizas'

SHEET_NETRAUKTI_VIDUTINIAM = 'netraukti vidutiniam'
SHEET_SANDELIAI = 'sandėliai'
SHEET_ATSARGOS = 'atsargos'
SHEET_OPERACIJOS = 'operacijos'
SHEET_AKCIZO_OPERACIJOS = "akcizo operacijos"
SHEET_NUOSTOLIAI = 'Nuostoliai'
SHEET_VNT_KONVERSIJA = 'vnt konversija'
SHEET_AKCIZO_SUMA = 'akcizo suma'
SHEET_LIKUTIS_MEN_PAB_AX = 'Likutis men pab Ax'

LIKUTIS_MEN_PRADZ = 'Likutis men pradz'
KS_VIENETAS = "KS vienetas"
KIEKIS = 'Kiekis'
AKCIZO_NUORODA = 'Akcizo nuoroda'
KODAS = "Kodas"
RINKA = "Rinka"
FIN_DATA = "Fin. data"
SANDELIS_2 = "Sandėlis 2"
SHEET_LIKUTIS_MEN_PRADZ = 'Likutis men pradz'
I_IS = "Į/Iš"
PARTIJOS_ID = "Partijos Id"
MENESIO_KIEKIS = 'Menesio kiekis'
MAX_KIEKIS = 'Max kiekis'
MENESIO_VID_AKCIZAS = 'Mėnesio vid. akcizas'
VIDUT_AKCIZAS_SUMA = 'Vidut. Akcizas Suma'
AKCIZO_TARIFAS = 'Akcizo tarifas'
PIRKIMAI = 'Pirkimai'
GAMYBA = 'Gamyba'
VIDUTINIS_AKCIZAS = 'Vidutinis akcizas'
LIKUTIS_DIENOS_PRADZIAI = 'Likutis dienos pradziai'
OPERACIJOS_VISAS = 'Operacijos visas'
KIEKIS_FINAL = 'Kiekis_final'
IMONE = 'Įmonė'
TARIFINE_GRUPE = 'Tarifinė grupė'
FAKTINE_DATA = 'Faktinė data'
SANDELIS = 'Sandėlis'
DAUGIKLIS = 'Daugiklis'
KOEFICIENTAS = 'Koeficientas'
VIENETAS = 'Vienetas'
STIPRUMAS = 'Stiprumas'
NUORODA = 'Nuoroda'
PREKES_NR = 'Prekės Nr.'
TRF_GR_KODAS = 'Tarifinės grupės kodas'
VNT_Y = 'Vienetas_y'
TO_VNT = 'Į vnt.'
VNT_X = 'Vienetas_x'
FROM_VNT = 'Iš vieneto'
VIETA = 'Vieta'
ISLAIDU_CENTRAS = 'Išlaidų centras'
SANDELIO_TIPAS = 'Sandėlio tipas'
NUOSTOLIO_TIPAS = 'Nuostolio tipas'
NUOSTOLIS_VISAS = 'Nuostolis visas'
SANDELIO_TIPAS_2 = SANDELIO_TIPAS + " 2"
TALPA = 'Talpa'
AKCIZO_SUMA = 'Akcizo suma'
AKCIZO_SUMA_FINAL = 'Akcizo suma final'
KIEKIS_FINAL_AX = 'Kiekis_final_AX'

# Stulpelių duomenys, kurie naudojami kaip konstantos
# VISOS AKCIZO NUORODOS
"""
    Netraukti
    Nuostolis
    Nuostolis-pilstymo
    KS eilutė
    KS
    Nuostolis-pilstymo
    Nuostolis-KOLLAB
    Netraukti
    Perkelti iš AAPS
    Perkelti į AAPS
    Pirkimo užsakymas
    Gamyba

"""
talpa = 'Talpa'  # Daugiklis
talpa_stipr = 'talpa * stiprumas'  # Daugiklis
gavimas = "Gavimas"  # Į/Iš
isdavimas = "Išdavimas"  # Į/Iš
perkelti = "Perkelti"  # Nuoroda


def save_to_pickle(df, name):
    if SAVE_TO_PICKLE:
        if not 'pickle' in name:
            name = name + ".pickle"
        df.to_pickle(name)


def calculate_final_qty(row):
    if row[VNT_X] == row[VNT_Y]:
        return round(row[KIEKIS], 3)
    else:
        if row[DAUGIKLIS].strip().lower() == talpa_stipr.strip().lower():
            kiekis = round(row[KIEKIS] * row[TALPA] * row[STIPRUMAS] / \
                           row[KOEFICIENTAS], 3)
            return kiekis
        elif row[DAUGIKLIS].strip().lower() == talpa.strip().lower():
            kiekis = round(row[KIEKIS] * row[TALPA] / row[KOEFICIENTAS], 3)
            return kiekis
        else:
            kiekis = round(row[KIEKIS], 3)
            return kiekis


def add_pirkimai(row):
    if row[NUORODA] == 'Pirkimo užsakymas':
        return row[KIEKIS_FINAL]
    else:
        return 0


def gamyba_be_suvest(row):
    if (row[NUORODA] == GAMYBA) and pd.isnull(row[GAMYBA]):
        return row[KIEKIS_FINAL]
    else:
        return 0


# def get_all_damages(row):
#     if row[SANDELIO_TIPAS] == nuostolio_sandėlis and row[NUORODA] == "Pardavimo užsakymas":
#         return row[KIEKIS_FINAL]
#     else:
#         return 0


# def get_all_gamybos_nuostolis(row):
#     if row[NUOSTOLIO_TIPAS] == gamybos_nuostolis and row[NUORODA] == "Pardavimo užsakymas":
#         return row[KIEKIS_FINAL]


# def get_all_saugojimo_nuostolis(row):
#     if row[NUOSTOLIO_TIPAS] == saugojimo_nuostolis and row[NUORODA] == "Pardavimo užsakymas":
#         return row[KIEKIS_FINAL]


# def get_all_virsnorminis_nuostolis(row):
#     if row[NUOSTOLIO_TIPAS] == virsnorminis_nuostolis and row[NUORODA] == "Pardavimo užsakymas":
#         return row[KIEKIS_FINAL]


# def zero_damage_warehouse(row):
#     if row[SANDELIO_TIPAS] == nuostolio_sandėlis:
#         return 0
#     else:
#         return row[KIEKIS_FINAL]


def get_pritrauktas_df(filename):
    global pritrauktas_df
    print('Pritraukiami duomenys...')

    ats_op_df = pd.read_excel(filename, sheet_name=SHEET_OPERACIJOS)
    atsargos_df = pd.read_excel(filename, sheet_name=SHEET_ATSARGOS)
    sandeliai_df = pd.read_excel(filename, sheet_name=SHEET_SANDELIAI)
    akcizo_operacijos_df = pd.read_excel(filename, sheet_name=SHEET_AKCIZO_OPERACIJOS)
    likutis_men_pr_df = pd.read_excel(filename, sheet_name=LIKUTIS_MEN_PRADZ)
    akcizo_suma = pd.read_excel(filename, sheet_name=SHEET_AKCIZO_SUMA)
    likutis_men_pr_df[KIEKIS].fillna(0, inplace=True)
    netraukti_vidutiniam_df = pd.read_excel(filename, sheet_name=SHEET_NETRAUKTI_VIDUTINIAM)
    netraukti_vidutiniam_df[GAMYBA] = 0

    atsargos_df.rename(columns={KS_VIENETAS: VIENETAS}, inplace=True)
    netraukti_vidutiniam_df.rename(columns={KS_VIENETAS: VIENETAS}, inplace=True)

    vnt_konv = pd.read_excel(filename, sheet_name=SHEET_VNT_KONVERSIJA)
    vnt_konv.rename(columns={FROM_VNT: VNT_X, TO_VNT: VNT_Y}, inplace=True)
    print("Viso ats_operacijų: ", len(ats_op_df))

    pritrauktas_df = pd.merge(ats_op_df, atsargos_df[[PREKES_NR, TARIFINE_GRUPE, TALPA, STIPRUMAS, VIENETAS]],
                              on=PREKES_NR)

    print(pritrauktas_df.columns)

    print("Pritraukta tarifinė grupė, talpa, stiprumas, vienetas. Eilučių skaičius: ", len(pritrauktas_df))
    pritrauktas_df = pd.merge(pritrauktas_df, tarif_group_df[[TARIFINE_GRUPE, VIENETAS, AKCIZO_TARIFAS]],
                              on=TARIFINE_GRUPE)

    print("Pritraukas vienetas: ", len(pritrauktas_df))
    pritrauktas_df = pd.merge(pritrauktas_df, vnt_konv[[KOEFICIENTAS, DAUGIKLIS, VNT_X, VNT_Y]],
                              on=[VNT_X, VNT_Y], how='left')

    print("Pritrauktas koeficientas, daugiklis, vnt_x, vnt_y : ", len(pritrauktas_df))
    pritrauktas_df = pd.merge(pritrauktas_df,
                              sandeliai_df[[SANDELIS, IMONE, SANDELIO_TIPAS]].drop_duplicates(subset=[SANDELIS]),
                              on=SANDELIS, how='left')

    print("Pritrauktas sandelis, sandelio tipas, imone:", len(pritrauktas_df))

    pritrauktas_df = pd.merge(pritrauktas_df, sandeliai_df[[SANDELIS, NUOSTOLIO_TIPAS, VIETA]], on=[SANDELIS, VIETA],
                              how='left')

    print("Nuostolio tipas: ", len(pritrauktas_df))
    pritrauktas_df[KIEKIS_FINAL] = pritrauktas_df.apply(calculate_final_qty, axis=1)
    pritrauktas_df[PIRKIMAI] = pritrauktas_df.apply(add_pirkimai, axis=1)

    pritrauktas_df = pd.merge(pritrauktas_df, netraukti_vidutiniam_df[[GAMYBA, PREKES_NR]], on=PREKES_NR,
                              how='left')
    print("GAMYBA, PREKES NR: ", len(pritrauktas_df))
    pritrauktas_df[GAMYBA] = pritrauktas_df.apply(gamyba_be_suvest, axis=1)

    # pritrauktas_df[NUOSTOLIS_VISAS] = pritrauktas_df.apply(get_all_damages, axis=1)
    # pritrauktas_df[NUOSTOLIS_GAMINANT] = pritrauktas_df.apply(get_all_gamybos_nuostolis, axis=1)
    # pritrauktas_df[NUOSTOLIS_SAUGANT] = pritrauktas_df.apply(get_all_saugojimo_nuostolis, axis=1)
    # pritrauktas_df[NUOSTOLIS_VIRSNORM] = pritrauktas_df.apply(get_all_virsnorminis_nuostolis, axis=1)

    ############### likutis AXAPTA#############
    likutis_men_pab_ax = pd.read_excel(filename, sheet_name=SHEET_LIKUTIS_MEN_PAB_AX)
    likutis_men_pab_ax = pd.merge(likutis_men_pab_ax, atsargos_df[[PREKES_NR, TARIFINE_GRUPE]],
                                  on=PREKES_NR)
    likutis_men_pab_ax = pd.merge(likutis_men_pab_ax, tarif_group_df[[TARIFINE_GRUPE, VIENETAS, AKCIZO_TARIFAS]],
                                  on=TARIFINE_GRUPE)
    likutis_men_pab_ax = pd.merge(likutis_men_pab_ax, vnt_konv[[KOEFICIENTAS, DAUGIKLIS, VNT_X, VNT_Y]],
                                  on=[VNT_X, VNT_Y], how='left')
    likutis_men_pab_ax = pd.merge(likutis_men_pab_ax,
                                  sandeliai_df[[SANDELIS, IMONE, SANDELIO_TIPAS]].drop_duplicates(subset=[SANDELIS]),
                                  on=SANDELIS, how='left')
    likutis_men_pab_ax[KIEKIS_FINAL_AX] = likutis_men_pab_ax.apply(calculate_final_qty, axis=1)

    likutis_men_pab_ax = likutis_men_pab_ax.groupby([IMONE, TARIFINE_GRUPE])[KIEKIS_FINAL_AX].sum()
    likutis_men_pab_ax = pd.DataFrame(likutis_men_pab_ax)
    save_to_pickle(likutis_men_pab_ax, 'lik_AX')
    ##########################################
    pritrauktas_df[SANDELIS_2] = pritrauktas_df.apply(get_sandelis2, axis=1)

    sandeliai_df.rename(
        columns={SANDELIS: SANDELIS_2,
                 SANDELIO_TIPAS: SANDELIO_TIPAS_2},
        inplace=True)

    pritrauktas_df = pd.merge(pritrauktas_df,
                              sandeliai_df[[SANDELIS_2, SANDELIO_TIPAS_2]].drop_duplicates(subset=[SANDELIS_2]),
                              on=SANDELIS_2,
                              how='left')
    print("SANDELIO TIPAS 2: ", len(pritrauktas_df))
    pritrauktas_df = pd.merge(pritrauktas_df, akcizo_operacijos_df,
                              on=[SANDELIO_TIPAS, NUORODA, I_IS, SANDELIO_TIPAS_2],
                              how='left')
    print("Akcizo operacijos: ", len(pritrauktas_df))
    pritrauktas_df[RINKA] = pritrauktas_df.apply(get_rinka, axis=1)

    likutis_men_pr_df.rename(columns={KIEKIS: LIKUTIS_MEN_PRADZ, SANDELIS: IMONE}, inplace=True)
    pritrauktas_df = pd.merge(pritrauktas_df, likutis_men_pr_df, on=[IMONE, TARIFINE_GRUPE], how='left')
    pritrauktas_df[ISLAIDU_CENTRAS] = pritrauktas_df[ISLAIDU_CENTRAS].astype(str)

    pritrauktas_df = pd.merge(pritrauktas_df, akcizo_suma[[NUORODA, ISLAIDU_CENTRAS, AKCIZO_SUMA]],
                              on=[NUORODA, ISLAIDU_CENTRAS], how='left')

    pritrauktas_df[AKCIZO_SUMA_FINAL] = np.where(pritrauktas_df[AKCIZO_SUMA].str.strip() == 'Taip',
                                                 pritrauktas_df[AKCIZO_TARIFAS] * pritrauktas_df[KIEKIS_FINAL],
                                                 0)
    pritrauktas_df[AKCIZO_SUMA_FINAL] = pritrauktas_df[AKCIZO_SUMA_FINAL].round(decimals=2)

    save_to_pickle(pritrauktas_df, 'pritrauktas_df')

    return (pritrauktas_df, likutis_men_pab_ax)


def get_final_df_grouped(pritrauktas_df):
    final_df = pritrauktas_df.groupby([IMONE, TARIFINE_GRUPE, FAKTINE_DATA])[
        [KIEKIS_FINAL, PIRKIMAI, GAMYBA]].sum()

    save_to_pickle(final_df, 'final_df')

    print("Duomenys pritraukti\nInicijuojamas vidutinio akcizo skaiciavimas...")

    return final_df


def add_likutis_men_pradziai(row):
    # if row.name[2] != pd.Timestamp(start_date):
    if row.name[2] != start_date:
        likutis_men_pr = 0
    else:
        likutis_men_pr = likutis_men_pr_df.loc[
            (likutis_men_pr_df[SANDELIS] == row.name[0]) & (likutis_men_pr_df[TARIFINE_GRUPE] == row.name[1])]
        likutis_men_pr = likutis_men_pr[KIEKIS].values[0]
    return likutis_men_pr


def get_sandelis2(row):
    if row[NUORODA] == perkelti:
        if row[I_IS] == isdavimas:
            ig = gavimas
        else:
            ig = isdavimas

        try:
            x = pritrauktas_df.loc[
                (pritrauktas_df[I_IS] == ig) &
                (pritrauktas_df[PARTIJOS_ID] == row[PARTIJOS_ID])].index[0]
            return pritrauktas_df.iloc[x][SANDELIS]
        except Exception as e:
            return "NERASTAS"
    else:
        return ""


def get_df_of_calc_avg(grouped_by_days_df):
    """Suskaičiuoja vidutinį likutį ir kitas mėnesiui reikalingas reikšmes

    :returns (grouped_by_days_df, monthly_report_df)"""
    idx = pd.date_range(start_date, end_date).date
    tarifai_all = likutis_men_pr_df[TARIFINE_GRUPE].unique()
    grouped_by_days_df = grouped_by_days_df \
        .unstack([IMONE, FAKTINE_DATA]) \
        .reindex(tarifai_all).fillna(0) \
        .stack([IMONE, FAKTINE_DATA]) \
        .unstack([IMONE, TARIFINE_GRUPE]) \
        .reindex(idx).fillna(0) \
        .stack([IMONE, TARIFINE_GRUPE]) \
        .swaplevel(0, 2) \
        .swaplevel(0, 1) \
        .groupby(level=[0, 1, 2]).sum()
    grouped_by_days_df.index.set_names(FAKTINE_DATA, level=2, inplace=True)
    grouped_by_days_df = pd.DataFrame(grouped_by_days_df)
    grouped_by_days_df.rename(columns={KIEKIS_FINAL: OPERACIJOS_VISAS}, inplace=True)
    grouped_by_days_df[LIKUTIS_DIENOS_PRADZIAI] = grouped_by_days_df.apply(add_likutis_men_pradziai, axis=1)

    grouped_by_days_df[VIDUTINIS_AKCIZAS] = np.nan
    grouped_by_days_df[VIDUT_AKCIZAS_SUMA] = np.nan
    grouped_by_days_df[MENESIO_KIEKIS] = np.nan
    grouped_by_days_df[AKCIZO_TARIFAS] = np.nan
    grouped_by_days_df[MENESIO_VID_AKCIZAS] = np.nan
    grouped_by_days_df[MAX_KIEKIS] = np.nan

    warehouse_level = grouped_by_days_df.index.levels[0]
    tarif_group_level = grouped_by_days_df.index.levels[1]
    month_days_level = grouped_by_days_df.index.levels[2]
    report_dict = {}
    for warehouse in warehouse_level:
        print("Skaičiuojamas sandėlis {0}".format(warehouse))
        for tarif in tarif_group_level:
            for i, (idx, row) in zip(np.arange(len(grouped_by_days_df.loc[warehouse, tarif].index)),
                                     grouped_by_days_df.loc[warehouse, tarif].iterrows()):
                row[VIDUTINIS_AKCIZAS] = row[LIKUTIS_DIENOS_PRADZIAI] + row[GAMYBA] + row[PIRKIMAI]
                likutis_kitos_dienos_pradziai = row[LIKUTIS_DIENOS_PRADZIAI] + row[OPERACIJOS_VISAS]
                try:
                    next_date = grouped_by_days_df.loc[warehouse, tarif].iloc[[i + 1]].index[0]
                    # grouped_by_days_df.loc[warehouse, tarif, next_date][
                    #     LIKUTIS_DIENOS_PRADZIAI] = likutis_kitos_dienos_pradziai
                    grouped_by_days_df.loc[
                        (warehouse, tarif, next_date), LIKUTIS_DIENOS_PRADZIAI] = likutis_kitos_dienos_pradziai
                except Exception as e:
                    pass

            try:
                akcizo_tarif = tarif_group_df.loc[tarif_group_df[TARIFINE_GRUPE] == tarif][AKCIZO_TARIFAS]
                # first_row = grouped_by_days_df.loc[warehouse, tarif, pd.Timestamp(start_date)]
                first_row = grouped_by_days_df.loc[warehouse, tarif, start_date]
                first_row[AKCIZO_TARIFAS] = akcizo_tarif.values[0]
                first_row[VIDUT_AKCIZAS_SUMA] = grouped_by_days_df.loc[warehouse, tarif][VIDUTINIS_AKCIZAS].sum().round(
                    2)
                first_row[MENESIO_KIEKIS] = (first_row[VIDUT_AKCIZAS_SUMA] / last_day).round(2)
                first_row[MENESIO_VID_AKCIZAS] = (first_row[MENESIO_KIEKIS] * first_row[AKCIZO_TARIFAS]).round(2)
                first_row[MAX_KIEKIS] = grouped_by_days_df.loc[warehouse, tarif][VIDUTINIS_AKCIZAS].max()
                # report_list.append([warehouse,
                #                     tarif,
                #                     first_row[AKCIZO_TARIFAS],
                #                     first_row[VIDUT_AKCIZAS_SUMA],
                #                     first_row[MENESIO_KIEKIS],
                #                     first_row[MENESIO_VID_AKCIZAS],
                #                     first_row[MAX_KIEKIS]
                #                     ])

                report_dict[(warehouse, tarif)] = [first_row[AKCIZO_TARIFAS],
                                                   first_row[VIDUT_AKCIZAS_SUMA],
                                                   first_row[MENESIO_KIEKIS],
                                                   first_row[MENESIO_VID_AKCIZAS],
                                                   first_row[MAX_KIEKIS]
                                                   ]

            except Exception as e:
                # print(e)
                pass
    monthly_report_df = pd.DataFrame(report_dict).T
    monthly_report_df.rename(columns={0: AKCIZO_TARIFAS,
                                      1: VIDUT_AKCIZAS_SUMA,
                                      2: MENESIO_KIEKIS,
                                      3: MENESIO_VID_AKCIZAS,
                                      4: MAX_KIEKIS}, inplace=True)

    print(monthly_report_df)

    # monthly_report_df.set_index(IMONE, inplace=True)
    return (grouped_by_days_df, monthly_report_df)


def format_all_sheets(writer):
    for sheet_name in writer.sheets:
        sheet = writer.sheets[sheet_name]
        max_col = sheet.max_column
        max_row = sheet.max_row
        cell_array = "A1:" \
                     + get_column_letter(max_col) \
                     + str(max_row)
        sheet.auto_filter.ref = cell_array
        sheet.freeze_panes = sheet['A2']
        if sheet_name == SHEET_AKCIZO_OPERACIJOS:
            sheet.freeze_panes = sheet["E3"]
            sheet.sheet_properties.tabColor = "069b4e"
            sheet.row_dimensions[2].height = 38
            sheet.row_dimensions[3].height = 48

        if sheet_name == SHEET_SUVESTINE:
            sheet.sheet_properties.tabColor = "069b4e"

        sheet.row_dimensions[1].height = 58
        for i in range(1, max_col + 1):
            width = 0
            sheet.cell(row=1, column=i).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for r in range(1, max_row + 1):
                try:
                    cell_length = len(sheet.cell(row=r, column=i).value)
                    if width < cell_length:
                        width = cell_length
                except TypeError:
                    pass
            if sheet_name == SHEET_AKCIZO_OPERACIJOS:
                width = 8.5
                sheet.cell(row=2, column=i).alignment = Alignment(horizontal='center', vertical='center',
                                                                  wrap_text=True)
                sheet.cell(row=3, column=i).alignment = Alignment(horizontal='center', vertical='center',
                                                                  wrap_text=True)

            if sheet_name == SHEET_SUVESTINE and i < 4:
                width = 15
            sheet.column_dimensions[get_column_letter(i)].width = width * 1.2


def run_calculation_from_ui(filename, save_file_name, update_text: tk.StringVar):
    try:
        run_calculation(filename, save_file_name, update_text)
    except Exception as e:
        update_text.set(e)


def get_rinka(row):
    try:
        return row[KODAS][0]
    except Exception:
        return ""


def pivot_frames(pritrauktas_df, likutis_men_pab_ax):
    pritrauktas_df["VISO"] = "VISO"  # fake, pivotus graziau sudėtų

    pivot_df = pd.pivot_table(pritrauktas_df, index=[IMONE, TARIFINE_GRUPE, VNT_Y, LIKUTIS_MEN_PRADZ],
                              columns=[AKCIZO_NUORODA, ISLAIDU_CENTRAS], values=KIEKIS_FINAL, aggfunc=np.sum)
    pivot_df_1 = pd.pivot_table(pritrauktas_df, index=[IMONE, TARIFINE_GRUPE, VNT_Y, LIKUTIS_MEN_PRADZ],
                                columns=[AKCIZO_NUORODA, "VISO"], values=KIEKIS_FINAL, aggfunc=np.sum)
    pivot_df_2 = pd.pivot_table(pritrauktas_df, index=[IMONE, TARIFINE_GRUPE, VNT_Y, LIKUTIS_MEN_PRADZ],
                                columns=[AKCIZO_NUORODA, RINKA], values=KIEKIS_FINAL, aggfunc=np.sum)
    pivot_df_3 = pd.pivot_table(pritrauktas_df, index=[IMONE, TARIFINE_GRUPE, VNT_Y, LIKUTIS_MEN_PRADZ],
                                columns=[AKCIZO_NUORODA, NUOSTOLIO_TIPAS], values=KIEKIS_FINAL, aggfunc=np.sum)

    pivot_df_4 = pd.pivot_table(pritrauktas_df, index=[IMONE, TARIFINE_GRUPE, VNT_Y, LIKUTIS_MEN_PRADZ],
                                columns=[AKCIZO_NUORODA, ISLAIDU_CENTRAS], values=AKCIZO_SUMA_FINAL, aggfunc=np.sum)
    pivot_df_5 = pd.pivot_table(pritrauktas_df, index=[IMONE, TARIFINE_GRUPE, VNT_Y, LIKUTIS_MEN_PRADZ],
                                columns=[AKCIZO_NUORODA, "VISO"], values=AKCIZO_SUMA_FINAL, aggfunc=np.sum)

    pivot_df_4.rename(columns={'Pardavimo užsakymas': 'Pardavimo užsakymas Akcizo suma',
                               'Pirkimo užsakymas': 'Pirkimo užsakymas Akcizo suma'}, level=0, inplace=True)
    pivot_df_5.rename(columns={'Pardavimo užsakymas': 'Pardavimo užsakymas Akcizo suma',
                               'Pirkimo užsakymas': 'Pirkimo užsakymas Akcizo suma'}, level=0, inplace=True)

    # df = pd.concat([pivot_df[[GAMYBA]],
    #                 pivot_df_1.loc[:, : "Netraukti"],
    #                 pivot_df_3[['Nuostolis']],
    #                 pivot_df_1.loc[:, "Nuostolis": "Nuostolis-pilstymo"],
    #                 pivot_df_1.loc[:, "Perkelti iš AAPS": "Perkelti į AAPS"],
    #                 pivot_df[['Pardavimo užsakymas']],
    #                 pivot_df_1[['Pardavimo užsakymas']],
    #                 pivot_df_4[['Pardavimo užsakymas Akcizo suma']],
    #                 pivot_df_5[['Pardavimo užsakymas Akcizo suma']],
    #                 pivot_df_2[['Pirkimo užsakymas']],
    #                 pivot_df_1[['Pirkimo užsakymas']],
    #                 pivot_df_4[['Pirkimo užsakymas Akcizo suma']],
    #                 pivot_df_5[['Pirkimo užsakymas Akcizo suma']]
    #                 ], axis=1)

    df = pd.concat([pivot_df[[GAMYBA]],
                    pivot_df_1.loc[:, : "Netraukti"],
                    pivot_df_3[['Nuostolis']],
                    pivot_df_1.loc[:, "Nuostolis": "Nuostolis-pilstymo"],
                    pivot_df_1.loc[:, "Perkelti iš AAPS": "Perkelti į AAPS"],
                    pivot_df[['Pardavimo užsakymas']],
                    pivot_df_1[['Pardavimo užsakymas']]], axis=1)
    try:
        df = pd.concat([df,
                        pivot_df_4[['Pardavimo užsakymas Akcizo suma']],
                        pivot_df_5[['Pardavimo užsakymas Akcizo suma']]], axis=1
                       )
    except Exception as e:
        print(e)

    df = pd.concat([df,
                    pivot_df_2[['Pirkimo užsakymas']],
                    pivot_df_1[['Pirkimo užsakymas']]], axis=1
                   )
    try:
        df = pd.concat([df,
                        pivot_df_4[['Pirkimo užsakymas Akcizo suma']],
                        pivot_df_5[['Pirkimo užsakymas Akcizo suma']]], axis=1
                       )
    except Exception as e:
        print(e)

    df['TOTAL'] = np.nan
    # df['TOTAL LIKUTIS PAB AX'] = np.nan
    likutis_men_pab_ax['AKCIZO OPERACIJOS TOTAL'] = np.nan
    likutis_men_pab_ax = likutis_men_pab_ax.reset_index()
    likutis_men_pab_ax.set_index([IMONE], inplace=True)
    # likutis_men_pab_ax[TARIFINE_GRUPE] = likutis_men_pab_ax[TARIFINE_GRUPE].astype(str)
    for i, row in pivot_df_1.iterrows():
        print(i)
        total = round(i[3] + row.sum(), 2)
        df.loc[i, 'TOTAL'] = total
        # likutis = round(likutis_men_pab_ax.loc[i[0], i[1]], 2)
        try:
            likutis_men_pab_ax.loc[(likutis_men_pab_ax.index == str(i[0])) & (likutis_men_pab_ax[TARIFINE_GRUPE] == i[1]), ['AKCIZO OPERACIJOS TOTAL']] = total
        except KeyError as e:
            print("error", str(e))
    # likutis_men_pab_ax.to_excel("test_lik_pab.xlsx")
    df = df.round(2)
    df.columns.set_names("Isl.centr / VISO / Nuos.tipas", level=1, inplace=True)
    save_to_pickle(pivot_df, 'pivot')
    save_to_pickle(pivot_df_1, 'pivot1')
    save_to_pickle(pivot_df_2, 'pivot2')
    save_to_pickle(df, 'concat_pivot')
    return (df, likutis_men_pab_ax)


def update_text(text, update_text: tk.StringVar):
    print(text)
    if update_text is not None and isinstance(update_text, tk.StringVar):
        update_text.set(text)


def run_calculation(filename, save_file_name, message):
    global tarif_group_df, last_day, start_date, end_date, likutis_men_pr_df
    start_time = time.time()

    error_messages = ""

    tarif_group_df = pd.read_excel(filename, sheet_name=SHEET_TARIFINES_GRUPES)
    tarif_group_df.rename(columns={TRF_GR_KODAS: TARIFINE_GRUPE}, inplace=True)
    update_text("Pritraukiami duomenys...", message)

    pritrauktasAndLikAx_dataframes = get_pritrauktas_df(filename)
    pritrauktas_df = pritrauktasAndLikAx_dataframes[0]
    likutis_men_pab_ax = pritrauktasAndLikAx_dataframes[1]

    random_time_stamp = pritrauktas_df[FAKTINE_DATA][randint(0, len(pritrauktas_df[FAKTINE_DATA]))]
    year = random_time_stamp.year
    month = random_time_stamp.month
    last_day = monthrange(year, month)[1]
    start_date = datetime.date(year, month, 1)
    end_date = datetime.date(year, month, last_day)
    if save_file_name is None:
        save_file_name = 'Vidutinis akcizas ' + str(year) + '-' + str(month)

    writer = pd.ExcelWriter(save_file_name + ".xlsx", engine='openpyxl')
    update_text("Grupuojama...", message)

    # dmg_df = pritrauktas_df.groupby([IMONE, TARIFINE_GRUPE])[
    #     [NUOSTOLIS_VISAS, NUOSTOLIS_GAMINANT, NUOSTOLIS_SAUGANT, NUOSTOLIS_VIRSNORM]].sum()

    # dmg_df.to_excel(writer, sheet_name=SHEET_NUOSTOLIAI)

    final_df = get_final_df_grouped(pritrauktas_df)
    likutis_men_pr_df = pd.read_excel(filename, sheet_name=SHEET_LIKUTIS_MEN_PRADZ)
    update_text("Skaičiuojamas vidutinis akcizas...", message)
    final_df_and_report = get_df_of_calc_avg(final_df)

    # frames = pivot_frames(pritrauktas_df, likutis_men_pab_ax)
    # pivot_df = frames[0]
    # likutis_men_pab_ax = frames[1]

    try:
        frames = pivot_frames(pritrauktas_df, likutis_men_pab_ax)
        pivot_df = frames[0]
        likutis_men_pab_ax = frames[1]
    except Exception as e:
        update_text(str(e), message)
        error_messages += str(e) + "\n"

    # pivot_df = pivot_frames(pritrauktas_df, likutis_men_pab_ax)

    update_text("Saugojama...", message)
    final_df_and_report[0].to_excel(writer, sheet_name=SHEET_AKCIZAS)
    final_df_and_report[1].to_excel(writer, sheet_name=SHEET_SUVESTINE)

    pritrauktas_df[FAKTINE_DATA] = pritrauktas_df[FAKTINE_DATA].dt.date
    pritrauktas_df[FIN_DATA] = pritrauktas_df[FIN_DATA].dt.date
    pritrauktas_df.to_excel(writer, sheet_name=SHEET_PRITRAUKTI_DUOMENYS)
    likutis_men_pab_ax.to_excel(writer, sheet_name=SHEET_LIKUTIS_PABAIGAI_AX_GRUPUOTAS)
    try:
        pivot_df.to_excel(writer, sheet_name=SHEET_AKCIZO_OPERACIJOS)
    except Exception as e:
        update_text(str(e), message)
        error_messages += str(e) + "\n"

    try:
        format_all_sheets(writer)
    except Exception as e:
        update_text(str(e), message)
    writer.save()
    update_text(error_messages + "Baigta\nVykdymo laikas: {0}s".format(int(time.time() - start_time)), message)


run_calculation("Duomenys akcizo deklaracijai MVGP+GUB 2019-02.xlsx", "Gubernija", None)